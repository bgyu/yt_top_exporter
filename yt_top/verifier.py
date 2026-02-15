import csv



def verify_csv(path: str) -> bool:
    required = ["category", "rank", "title", "channel", "views", "url", "published_at"]
    with open(path, newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        # check headers
        for h in required:
            if h not in r.fieldnames:
                return False
        rows = list(r)
        if len(rows) < 1:
            return False
        # check urls
        for row in rows:
            if not row.get("url", "").startswith("http"):
                return False
    return True


def verify_xlsx(path: str) -> bool:
    try:
        from openpyxl import load_workbook
    except Exception:
        # openpyxl not installed — fallback: attempt to read worksheet XML from the .xlsx zip
        try:
            import zipfile
            with zipfile.ZipFile(path) as z:
                try:
                    data = z.read('xl/worksheets/sheet1.xml').decode('utf-8', errors='ignore')
                except Exception:
                    # fall back to raw text read
                    with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                        data = f.read()
        except Exception:
            return False
        if 'url' not in data:
            return False
        if 'http' in data:
            return True
        return False

    wb = load_workbook(path)
    ws = wb.active
    # find URL column index by header
    headers = [c.value for c in next(ws.rows)]
    try:
        url_idx = headers.index("url")
    except ValueError:
        return False
    # check at least one hyperlink present in url column
    for row in ws.iter_rows(min_row=2, min_col=url_idx + 1, max_col=url_idx + 1):
        cell = row[0]
        if getattr(cell, "hyperlink", None):
            return True
    return False


def verify_all(csv_path: str, xlsx_path: str) -> bool:
    return verify_csv(csv_path) and verify_xlsx(xlsx_path)
