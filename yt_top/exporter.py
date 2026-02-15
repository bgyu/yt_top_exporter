import csv
import os
import time
from datetime import datetime, timedelta
from typing import List

import requests
from datetime import timezone

# Defaults and small config
DEFAULT_MAX_PER_CATEGORY = 10
DEFAULT_REGION = "US"
OUTPUT_CSV_DEFAULT = "youtube_top_videos_last_{days}_{lang}.csv"
ALLOWED_LANG_PREFIX = ("en", "zh")


def _mock_videos(category: str, n: int):
    now = datetime.utcnow().isoformat() + "Z"
    items = []
    for i in range(1, n + 1):
        items.append(
            {
                "category": category,
                "rank": i,
                "title": f"Mock Video {i} ({category})",
                "channel": "MockChannel",
                "views": 1000 * i,
                "url": f"http://example.com/{category}/{i}",
                "published_at": now,
            }
        )
    return items


def fetch_videos_for_category(category: str, n: int, lang: str, days: int, api_key: str):
    # Use YouTube Data API v3 'videos?chart=mostPopular' to fetch popular videos
    params = {
        "part": "snippet,statistics",
        "chart": "mostPopular",
        "regionCode": lang,
        "maxResults": min(50, n),
        "key": api_key,
    }
    # If category looks numeric, use as videoCategoryId; otherwise skip filtering
    if category and category.isdigit():
        params["videoCategoryId"] = category

    resp = requests.get("https://www.googleapis.com/youtube/v3/videos", params=params, timeout=10)
    resp.raise_for_status()
    data = resp.json()
    items = []
    for idx, it in enumerate(data.get("items", [])[:n], start=1):
        snip = it.get("snippet", {})
        stats = it.get("statistics", {})
        vid_id = it.get("id")
        url = f"https://www.youtube.com/watch?v={vid_id}"
        items.append(
            {
                "category": category,
                "rank": idx,
                "title": snip.get("title"),
                "channel": snip.get("channelTitle"),
                "views": int(stats.get("viewCount", 0)),
                "url": url,
                "published_at": snip.get("publishedAt"),
            }
        )
    return items


def write_csv(path: str, rows: List[dict]):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    headers = ["category", "rank", "title", "channel", "views", "url", "published_at"]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        for r in rows:
            w.writerow(r)

def write_xlsx_minimal(path: str, rows: List[dict]):
    """Write a minimal valid XLSX (ZIP+XML) including hyperlinks so Excel can open it.

    This produces a very small workbook with one sheet.
    """
    import zipfile
    from xml.sax.saxutils import escape

    headers = ["category", "rank", "title", "channel", "views", "url", "published_at"]

    def col_letter(i: int) -> str:
        s = ""
        while i > 0:
            i, rem = divmod(i - 1, 26)
            s = chr(65 + rem) + s
        return s

    def cell_xml(col: str, row_idx: int, value: object) -> str:
        v = "" if value is None else escape(str(value))
        return f'<c r="{col}{row_idx}" t="inlineStr"><is><t>{v}</t></is></c>'

    rows_xml = []
    # header
    header_cells = [cell_xml(col_letter(i), 1, h) for i, h in enumerate(headers, start=1)]
    rows_xml.append(f'<row r="1">{"".join(header_cells)}</row>')

    hyperlinks = []
    for ridx, r in enumerate(rows, start=2):
        cells = []
        for i, h in enumerate(headers, start=1):
            val = r.get(h, "")
            col = col_letter(i)
            cells.append(cell_xml(col, ridx, val))
            if h == "url" and isinstance(val, str) and val.startswith("http"):
                hyperlinks.append((f"{col}{ridx}", val))
        rows_xml.append(f'<row r="{ridx}">{"".join(cells)}</row>')

    hyperlinks_xml = "" if not hyperlinks else f"<hyperlinks>{''.join(f'<hyperlink ref="{cell}" r:id="rId{idx+1}"/>' for idx, (cell, _) in enumerate(hyperlinks))}</hyperlinks>"

    sheet_xml = f"""<?xml version='1.0' encoding='UTF-8'?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheetData>
    {''.join(rows_xml)}
  </sheetData>
  {hyperlinks_xml}
</worksheet>"""

    # content types and relationships
    content_types = """<?xml version='1.0' encoding='UTF-8'?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
</Types>"""

    rels_xml = """<?xml version='1.0' encoding='UTF-8'?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="/xl/workbook.xml"/>
</Relationships>"""

    workbook_xml = """<?xml version='1.0' encoding='UTF-8'?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets>
    <sheet name="Sheet1" sheetId="1" r:id="rId1"/>
  </sheets>
</workbook>"""

    workbook_rels = """<?xml version='1.0' encoding='UTF-8'?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
</Relationships>"""

    sheet_rels = ""
    if hyperlinks:
        rels = []
        for idx, (_, target) in enumerate(hyperlinks, start=1):
            rid = f"rId{idx+1}"
            # Target must be XML-escaped
            rels.append(f'<Relationship Id="{rid}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink" Target="{escape(target)}" TargetMode="External"/>')
        sheet_rels = f"<?xml version='1.0' encoding='UTF-8'?><Relationships xmlns='http://schemas.openxmlformats.org/package/2006/relationships'>{''.join(rels)}</Relationships>"

    # write zip package
    with zipfile.ZipFile(path, 'w', compression=zipfile.ZIP_DEFLATED) as z:
        z.writestr('[Content_Types].xml', content_types)
        z.writestr('_rels/.rels', rels_xml)
        z.writestr('xl/workbook.xml', workbook_xml)
        z.writestr('xl/_rels/workbook.xml.rels', workbook_rels)
        z.writestr('xl/worksheets/sheet1.xml', sheet_xml)
        if sheet_rels:
            z.writestr('xl/worksheets/_rels/sheet1.xml.rels', sheet_rels)


def write_xlsx(path: str, rows: List[dict]):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except Exception:
        # fallback to minimal XLSX writer
        write_xlsx_minimal(path, rows)
        return

    wb = Workbook()
    ws = wb.active
    headers = ["category", "rank", "title", "channel", "views", "url", "published_at"]
    ws.append(headers)
    for r in rows:
        row = [r.get(h) for h in headers]
        ws.append(row)
        # set hyperlink on the last appended row's URL cell and ensure cell value is the URL
        url_idx = headers.index("url") + 1
        url_cell = ws.cell(row=ws.max_row, column=url_idx)
        url = r.get("url")
        if url:
            url_cell.value = url
            url_cell.hyperlink = url
            url_cell.font = Font(color="0000FF", underline="single")
    wb.save(path)
    return


def get_video_categories(api_key: str, region: str = "US"):
    """Return mapping of category id -> category title for the provided region."""
    params = {"part": "snippet", "regionCode": region, "key": api_key}
    resp = requests.get("https://www.googleapis.com/youtube/v3/videoCategories", params=params, timeout=10)
    resp.raise_for_status()
    data = resp.json()
    mapping = {}
    for it in data.get("items", []):
        cid = it.get("id")
        title = it.get("snippet", {}).get("title")
        mapping[cid] = title
    return mapping


def _detect_lang(text: str) -> str:
    """Very small heuristic to detect 'en' vs 'zh' based on characters."""
    if not text:
        return "en"
    for ch in text:
        # basic CJK range
        if "\u4e00" <= ch <= "\u9fff":
            return "zh"
    return "en"


def write_enriched_csv(path: str, rows: List[dict]):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    headers = [
        "category_id",
        "category_name",
        "title",
        "channel",
        "views",
        "language",
        "published_at",
        "video_url",
    ]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in rows:
            w.writerow([
                r.get("category_id", ""),
                r.get("category_name", ""),
                r.get("title", ""),
                r.get("channel", ""),
                r.get("views", ""),
                r.get("language", ""),
                r.get("published_at", ""),
                r.get("video_url", ""),
            ])


def fetch_and_export(categories: str, n: int, lang: str, days: int, api_key: str = None, mock: bool = False):
    cats = [c.strip() for c in categories.split(",")] if categories else ["all"]

    category_map = {}
    if not mock:
        if not api_key:
            raise RuntimeError("YOUTUBE_API_KEY is required when not running in mock mode")
        try:
            category_map = get_video_categories(api_key, region=lang)
        except Exception as e:
            raise RuntimeError(f"Failed to fetch video categories: {e}")

    # map category names to ids when possible
    mapped_cats = []
    for c in cats:
        if c.isdigit():
            mapped_cats.append(c)
            continue
        # try to find id by matching category title
        found = None
        for cid, title in category_map.items():
            if title and title.lower() == c.lower():
                found = cid
                break
        mapped_cats.append(found if found is not None else c)

    # if user requested "all", expand to all category ids we fetched
    if any((str(x).lower() == "all" for x in cats)) and category_map:
        mapped_cats = list(category_map.keys())

    all_rows = []
    for c in mapped_cats:
        if mock:
            rows = _mock_videos(c or "all", n)
        else:
            if not api_key:
                raise RuntimeError("YOUTUBE_API_KEY is required to fetch real data")
            # attempt to fetch; on HTTP errors skip this category but continue
            try:
                rows = fetch_videos_for_category(c, n, lang, days, api_key)
            except Exception as e:
                # log to stderr and skip this category
                try:
                    import sys

                    print(f"Skipping category {c}: {e}", file=sys.stderr)
                except Exception:
                    pass
                rows = []
        all_rows.extend(rows)

    out_csv = os.path.join("out", "top_videos.csv")
    out_xlsx = os.path.join("out", "top_videos.xlsx")
    write_csv(out_csv, all_rows)
    write_xlsx(out_xlsx, all_rows)

    # build enriched rows for the separate CSV
    enriched = []
    for r in all_rows:
        raw_cat = str(r.get("category", ""))
        category_id = ""
        category_name = ""
        if raw_cat.isdigit():
            category_id = raw_cat
            category_name = category_map.get(raw_cat, "")
        else:
            # maybe it's a name
            category_name = raw_cat
            # try to find id by name
            for cid, title in category_map.items():
                if title and title.lower() == raw_cat.lower():
                    category_id = cid
                    break

        title = r.get("title")
        enriched.append(
            {
                "category_id": category_id,
                "category_name": category_name,
                "title": title,
                "channel": r.get("channel"),
                "views": r.get("views"),
                "language": _detect_lang(title),
                "published_at": r.get("published_at"),
                "video_url": r.get("url"),
            }
        )
    # filter by allowed language prefixes
    filtered = [e for e in enriched if any(e.get("language", "").startswith(p) for p in ALLOWED_LANG_PREFIX)]

    enriched_csv = os.path.join("out", OUTPUT_CSV_DEFAULT.format(days=days, lang=lang))
    write_enriched_csv(enriched_csv, filtered)

    return out_csv, out_xlsx, enriched_csv
